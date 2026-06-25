import io
import json
import sqlite3
import urllib.parse
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# =========================================================
# APP CONFIG
# =========================================================
st.set_page_config(
    page_title="SPOTT Multi-Project E&I Tracker",
    page_icon="📊",
    layout="wide",
)

DB_FILE = "spott_project_tracker.db"

STATUS_COLOURS = {
    "OVERDUE": "#F8696B",       # red
    "DUE THIS WEEK": "#FFD966", # yellow
    "ON TRACK": "#C6E0B4",      # green
    "COMPLETE": "#A9D18E",
    "ORDERED": "#A9D18E",
    "NO DATE": "#D9EAD3",
    "READY": "#A9D18E",
    "PARTIAL": "#FFD966",
    "NOT READY": "#F8696B",
}

FAT_CRITICAL_DOC_KEYWORDS = [
    "Instrument Index", "System Architecture", "Blocks", "Junction Boxes", "Local Control Panel",
    "Control Narrative", "Cable Schedule", "Instrument Datasheets", "C&E", "Sequence Logic",
    "UCP WIRING", "UCP GA", "I/O List", "DCS List", "Electrical Load List", "UCP FAT Procedure",
]

FAT_CRITICAL_PROC_KEYWORDS = [
    "Bently", "Unit Control Panel", "Local Control", "Junction", "E&I Installation",
    "Compressor RTD", "Cables", "MDM", "Pressure", "Temperature", "Flow", "Level", "E-Stop",
]

DEFAULT_MILESTONES = [
    "Programming complete",
    "UCP inspection at vendor complete",
    "Internal UCP FAT setup complete",
    "Client UCP witness FAT ready",
    "Loop check readiness confirmed",
    "String test readiness confirmed",
]

DATE_COLUMNS = [
    "Expected Send to Client", "Current Forecast Send", "Actual Sent Date",
    "Expected Return from Client", "Current Forecast Return", "Expected RFQ Date",
    "Actual RFQ Date", "Expected PO/Order Date", "Actual PO/Order Date",
    "Expected Delivery Date", "Baseline Build/Test", "Start Date", "Complete Date",
]

# =========================================================
# GENERAL HELPERS
# =========================================================
def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        converted = pd.to_datetime(value, errors="coerce")
        if pd.notna(converted):
            return converted.date()
    except Exception:
        pass
    return None


def is_tick(value):
    return str(value).strip() in {"✓", "✔", "Y", "YES", "Yes", "yes", "Done", "DONE", "Complete", "Completed", "True", "true"}


def fmt_date(value):
    if value is None or pd.isna(value):
        return ""
    return pd.to_datetime(value).strftime("%d-%b-%Y")


def status_from_due(expected_date, done, today):
    if done:
        return "COMPLETE"
    if expected_date is None or pd.isna(expected_date):
        return "NO DATE"
    days_remaining = (expected_date - today).days
    if days_remaining < 0:
        return "OVERDUE"
    if days_remaining <= 7:
        return "DUE THIS WEEK"
    return "ON TRACK"


def overall_procurement_status(row, today):
    rfq_status = status_from_due(row.get("Expected RFQ Date"), bool(row.get("RFQ Placed?", False)), today)
    po_status = status_from_due(row.get("Expected PO/Order Date"), bool(row.get("Ordered?", False)), today)
    delivery_date = row.get("Expected Delivery Date")

    if rfq_status == "OVERDUE" or po_status == "OVERDUE":
        return "OVERDUE"
    if delivery_date and (delivery_date - today).days < 0:
        return "OVERDUE"
    if rfq_status == "DUE THIS WEEK" or po_status == "DUE THIS WEEK":
        return "DUE THIS WEEK"
    if delivery_date and 0 <= (delivery_date - today).days <= 7:
        return "DUE THIS WEEK"
    if row.get("RFQ Placed?", False) and row.get("Ordered?", False):
        return "ORDERED"
    return "ON TRACK"


def style_status(df, status_column):
    def row_style(row):
        colour = STATUS_COLOURS.get(str(row.get(status_column, "")), "#FFFFFF")
        return [f"background-color: {colour}"] * len(row)
    return df.style.apply(row_style, axis=1)


def readiness_status(score):
    if score >= 90:
        return "READY"
    if score >= 70:
        return "PARTIAL"
    return "NOT READY"


def mailto_link(email, subject, body):
    return "mailto:" + urllib.parse.quote(email or "") + "?subject=" + urllib.parse.quote(subject) + "&body=" + urllib.parse.quote(body)

# =========================================================
# DATABASE HELPERS
# =========================================================
def get_conn():
    return sqlite3.connect(DB_FILE, check_same_thread=False)


def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT NOT NULL UNIQUE,
            source_filename TEXT,
            docs_json TEXT,
            procurement_json TEXT,
            milestones_json TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)
    conn.commit()
    conn.close()


def df_to_json(df):
    if df is None or df.empty:
        return "[]"
    out = df.copy()
    for col in out.columns:
        if col in DATE_COLUMNS:
            out[col] = out[col].apply(lambda x: x.isoformat() if isinstance(x, date) else (pd.to_datetime(x).date().isoformat() if pd.notna(pd.to_datetime(x, errors="coerce")) else None))
    return out.to_json(orient="records")


def json_to_df(data):
    if not data:
        return pd.DataFrame()
    try:
        records = json.loads(data)
        df = pd.DataFrame(records)
        for col in df.columns:
            if col in DATE_COLUMNS:
                df[col] = df[col].apply(to_date)
        return df
    except Exception:
        return pd.DataFrame()


def save_project(project_name, source_filename, docs_df, proc_df, milestones_df):
    now = datetime.now().isoformat(timespec="seconds")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM projects WHERE project_name = ?", (project_name,))
    existing = cur.fetchone()
    if existing:
        cur.execute("""
            UPDATE projects
            SET source_filename = ?, docs_json = ?, procurement_json = ?, milestones_json = ?, updated_at = ?
            WHERE project_name = ?
        """, (source_filename, df_to_json(docs_df), df_to_json(proc_df), df_to_json(milestones_df), now, project_name))
    else:
        cur.execute("""
            INSERT INTO projects (project_name, source_filename, docs_json, procurement_json, milestones_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (project_name, source_filename, df_to_json(docs_df), df_to_json(proc_df), df_to_json(milestones_df), now, now))
    conn.commit()
    conn.close()


def list_projects():
    conn = get_conn()
    df = pd.read_sql_query("SELECT project_name, source_filename, created_at, updated_at FROM projects ORDER BY updated_at DESC", conn)
    conn.close()
    return df


def load_project(project_name):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT source_filename, docs_json, procurement_json, milestones_json FROM projects WHERE project_name = ?", (project_name,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return None, pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    source_filename, docs_json, proc_json, milestones_json = row
    return source_filename, json_to_df(docs_json), json_to_df(proc_json), json_to_df(milestones_json)


def delete_project(project_name):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM projects WHERE project_name = ?", (project_name,))
    conn.commit()
    conn.close()


def export_database_json():
    conn = get_conn()
    df = pd.read_sql_query("SELECT * FROM projects", conn)
    conn.close()
    return df.to_json(orient="records").encode("utf-8")


def import_database_json(uploaded_backup):
    records = json.loads(uploaded_backup.getvalue().decode("utf-8"))
    conn = get_conn()
    cur = conn.cursor()
    for r in records:
        cur.execute("""
            INSERT OR REPLACE INTO projects
            (id, project_name, source_filename, docs_json, procurement_json, milestones_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            r.get("id"), r.get("project_name"), r.get("source_filename"), r.get("docs_json"),
            r.get("procurement_json"), r.get("milestones_json"), r.get("created_at"), r.get("updated_at")
        ))
    conn.commit()
    conn.close()

# =========================================================
# SPOTT EXTRACTION
# =========================================================
def extract_project_name(file_bytes, fallback_name):
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        if "Internal Dates" in wb.sheetnames:
            val = wb["Internal Dates"]["F1"].value
            if val:
                return str(val).strip()
    except Exception:
        pass
    return Path(fallback_name).stem


def extract_document_schedule(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Internal Dates" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["Internal Dates"]
    project = ws["F1"].value or ""
    rows = []
    for r in range(5, ws.max_row + 1):
        doc_no = ws.cell(r, 1).value
        doc_title = ws.cell(r, 2).value
        if isinstance(doc_no, str) and doc_no.startswith("SCW"):
            expected_send = forecast_send = expected_return = forecast_return = None
            for rr in range(r + 1, min(r + 5, ws.max_row + 1)):
                label = str(ws.cell(rr, 1).value or "")
                if isinstance(ws.cell(rr, 1).value, str) and ws.cell(rr, 1).value.startswith("SCW"):
                    break
                if "Date Send" in label:
                    expected_send = to_date(ws.cell(rr, 3).value)
                    forecast_send = to_date(ws.cell(rr, 4).value)
                if "Date Returned" in label:
                    expected_return = to_date(ws.cell(rr, 3).value)
                    forecast_return = to_date(ws.cell(rr, 4).value)
            if expected_send or forecast_send or expected_return or forecast_return:
                rows.append({
                    "Project": project,
                    "Document No": doc_no,
                    "Document Title": doc_title or "",
                    "Expected Send to Client": expected_send,
                    "Current Forecast Send": forecast_send,
                    "Sent to Client?": False,
                    "Actual Sent Date": None,
                    "Expected Return from Client": expected_return,
                    "Current Forecast Return": forecast_return,
                    "Required for FAT?": any(k.lower() in str(doc_title).lower() for k in FAT_CRITICAL_DOC_KEYWORDS),
                    "Delay Reason / Notes": "",
                })
    return pd.DataFrame(rows)


def extract_procurement_schedule(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Internal Dates" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["Internal Dates"]
    project = ws["F1"].value or ""
    rows = []
    current_doc_no = ""
    current_doc_title = ""
    for r in range(5, ws.max_row + 1):
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value
        if isinstance(col_a, str) and col_a.startswith("SCW"):
            current_doc_no = col_a
            current_doc_title = col_b or ""
        item = ws.cell(r, 9).value
        rfq_due = to_date(ws.cell(r, 13).value)
        po_due = to_date(ws.cell(r, 14).value)
        delivery_due = to_date(ws.cell(r, 15).value)
        if item and (rfq_due or po_due or delivery_due):
            rows.append({
                "Project": project,
                "Item / Instrument": str(item).strip(),
                "Linked Document No": current_doc_no,
                "Linked Document Title": current_doc_title,
                "Expected RFQ Date": rfq_due,
                "RFQ Placed?": is_tick(ws.cell(r, 7).value),
                "Actual RFQ Date": None,
                "Expected PO/Order Date": po_due,
                "Ordered?": is_tick(ws.cell(r, 8).value),
                "Actual PO/Order Date": None,
                "Expected Delivery Date": delivery_due,
                "Received / Delivered?": False,
                "Baseline Build/Test": to_date(ws.cell(r, 16).value),
                "Required for FAT?": any(k.lower() in str(item).lower() for k in FAT_CRITICAL_PROC_KEYWORDS),
                "Dependency / Notes": ws.cell(r, 18).value or "",
                "Delay Reason / Notes": "",
            })
    return pd.DataFrame(rows)


def default_milestones():
    return pd.DataFrame([{
        "Milestone": m,
        "Start Date": None,
        "Complete Date": None,
        "Complete?": False,
        "Notes": "",
    } for m in DEFAULT_MILESTONES])

# =========================================================
# STATUS / READINESS
# =========================================================
def apply_document_status(df, today):
    out = df.copy()
    if out.empty:
        return out
    out["Document Status"] = out.apply(
        lambda row: status_from_due(row.get("Current Forecast Send") or row.get("Expected Send to Client"), bool(row.get("Sent to Client?", False)), today), axis=1
    )
    return out


def apply_procurement_status(df, today):
    out = df.copy()
    if out.empty:
        return out
    out["RFQ Status"] = out.apply(lambda row: status_from_due(row.get("Expected RFQ Date"), bool(row.get("RFQ Placed?", False)), today), axis=1)
    out["PO / Order Status"] = out.apply(lambda row: status_from_due(row.get("Expected PO/Order Date"), bool(row.get("Ordered?", False)), today), axis=1)
    out["Overall Status"] = out.apply(lambda row: overall_procurement_status(row, today), axis=1)
    return out


def calculate_fat_readiness(docs, procurement, milestones):
    fat_docs = docs[docs["Required for FAT?"] == True] if not docs.empty and "Required for FAT?" in docs.columns else docs
    fat_proc = procurement[procurement["Required for FAT?"] == True] if not procurement.empty and "Required for FAT?" in procurement.columns else procurement

    doc_total = len(fat_docs)
    doc_done = int(fat_docs["Sent to Client?"].fillna(False).astype(bool).sum()) if doc_total else 0
    doc_score = round((doc_done / doc_total) * 100, 1) if doc_total else 100.0

    proc_total = len(fat_proc)
    proc_done = int((fat_proc["Ordered?"].fillna(False).astype(bool) | fat_proc["Received / Delivered?"].fillna(False).astype(bool)).sum()) if proc_total else 0
    proc_score = round((proc_done / proc_total) * 100, 1) if proc_total else 100.0

    ms_total = len(milestones)
    ms_done = int(milestones["Complete?"].fillna(False).astype(bool).sum()) if ms_total and "Complete?" in milestones.columns else 0
    ms_score = round((ms_done / ms_total) * 100, 1) if ms_total else 100.0

    overall = round((doc_score * 0.35) + (proc_score * 0.45) + (ms_score * 0.20), 1)
    return {
        "Document Readiness %": doc_score,
        "Procurement Readiness %": proc_score,
        "FAT Milestone Readiness %": ms_score,
        "Overall FAT Readiness %": overall,
        "FAT Readiness Status": readiness_status(overall),
        "FAT Documents Complete": f"{doc_done}/{doc_total}",
        "FAT Procurement Complete": f"{proc_done}/{proc_total}",
        "FAT Milestones Complete": f"{ms_done}/{ms_total}",
    }


def project_summary(project_name, docs, proc, milestones, today):
    docs_s = apply_document_status(docs, today)
    proc_s = apply_procurement_status(proc, today)
    readiness = calculate_fat_readiness(docs_s, proc_s, milestones)
    return {
        "Project": project_name,
        "Documents": len(docs_s),
        "Procurement Items": len(proc_s),
        "Overdue Documents": int((docs_s.get("Document Status", pd.Series(dtype=str)) == "OVERDUE").sum()) if not docs_s.empty else 0,
        "Overdue Procurement": int((proc_s.get("Overall Status", pd.Series(dtype=str)) == "OVERDUE").sum()) if not proc_s.empty else 0,
        "FAT Readiness %": readiness["Overall FAT Readiness %"],
        "FAT Status": readiness["FAT Readiness Status"],
    }


def make_excel_download(docs, procurement, milestones, readiness_df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        readiness_df.to_excel(writer, sheet_name="FAT Readiness", index=False)
        docs.to_excel(writer, sheet_name="Documents", index=False)
        procurement.to_excel(writer, sheet_name="Procurement", index=False)
        milestones.to_excel(writer, sheet_name="FAT Milestones", index=False)
    output.seek(0)
    return output.getvalue()

# =========================================================
# APP START
# =========================================================
init_db()

st.title("📊 SPOTT Multi-Project E&I Tracker")
st.caption("Upload multiple SPOTT Excel files, save project changes, reload projects, and track Documents / Procurement / FAT readiness.")

st.warning(
    "Persistence note: this app saves project data into a SQLite database file. This is persistent for local/server deployments. "
    "On Streamlit Community Cloud, storage may reset when the app is redeployed or restarted. Use the Backup button regularly, or connect to a permanent database/SharePoint for production use."
)

assessment_date = st.sidebar.date_input("Assessment date", value=date.today())
engineer_name = st.sidebar.text_input("Engineer name", value="Engineer")
engineer_email = st.sidebar.text_input("Engineer email", value="", placeholder="name@company.com")

page = st.sidebar.radio(
    "Navigation",
    ["Project Library", "Portfolio Dashboard", "Project Dashboard", "Documents Tracker", "Procurement Tracker", "FAT Readiness", "Email Reminders", "Export / Backup"],
)

# =========================================================
# PROJECT LIBRARY
# =========================================================
if page == "Project Library":
    st.subheader("Project Library")
    st.markdown("Upload one or more SPOTT Excel files. Each file will be stored as a separate project.")

    uploads = st.file_uploader("Upload SPOTT Excel files", type=["xlsx", "xlsm"], accept_multiple_files=True)
    if st.button("Import uploaded Excel files", type="primary"):
        if not uploads:
            st.warning("Please upload at least one Excel file.")
        else:
            imported = []
            for f in uploads:
                file_bytes = f.getvalue()
                project_name = extract_project_name(file_bytes, f.name)
                docs = extract_document_schedule(file_bytes)
                proc = extract_procurement_schedule(file_bytes)
                milestones = default_milestones()
                save_project(project_name, f.name, docs, proc, milestones)
                imported.append(project_name)
            st.success("Imported projects: " + ", ".join(imported))

    st.markdown("### Stored projects")
    projects_df = list_projects()
    if projects_df.empty:
        st.info("No projects stored yet.")
    else:
        st.dataframe(projects_df, use_container_width=True, hide_index=True)

    st.markdown("### Delete a project")
    if not projects_df.empty:
        delete_name = st.selectbox("Select project to delete", projects_df["project_name"].tolist())
        if st.button("Delete selected project"):
            delete_project(delete_name)
            st.success(f"Deleted project: {delete_name}")
            st.rerun()

# =========================================================
# PORTFOLIO DASHBOARD
# =========================================================
elif page == "Portfolio Dashboard":
    st.subheader("Portfolio Dashboard")
    projects_df = list_projects()
    if projects_df.empty:
        st.info("No projects stored yet. Go to Project Library and upload SPOTT files.")
    else:
        summaries = []
        for project_name in projects_df["project_name"]:
            _, docs, proc, milestones = load_project(project_name)
            summaries.append(project_summary(project_name, docs, proc, milestones, assessment_date))
        summary_df = pd.DataFrame(summaries)
        st.dataframe(summary_df, use_container_width=True, hide_index=True)
        st.markdown("### FAT readiness by project")
        st.bar_chart(summary_df.set_index("Project")["FAT Readiness %"])

# =========================================================
# LOAD SELECTED PROJECT FOR PROJECT PAGES
# =========================================================
def selected_project_or_stop():
    projects_df = list_projects()
    if projects_df.empty:
        st.info("No projects stored yet. Go to Project Library and import SPOTT files first.")
        st.stop()
    return st.sidebar.selectbox("Select project", projects_df["project_name"].tolist())

if page in ["Project Dashboard", "Documents Tracker", "Procurement Tracker", "FAT Readiness", "Email Reminders", "Export / Backup"]:
    project_name = selected_project_or_stop()
    source_filename, docs_df, proc_df, milestones_df = load_project(project_name)
    if "current_project" not in st.session_state or st.session_state.current_project != project_name:
        st.session_state.current_project = project_name
        st.session_state.docs_edit = docs_df
        st.session_state.proc_edit = proc_df
        st.session_state.milestones_edit = milestones_df if not milestones_df.empty else default_milestones()

    docs_scored = apply_document_status(st.session_state.docs_edit, assessment_date)
    proc_scored = apply_procurement_status(st.session_state.proc_edit, assessment_date)
    readiness = calculate_fat_readiness(docs_scored, proc_scored, st.session_state.milestones_edit)
    readiness_df = pd.DataFrame([readiness])

    def save_current_project():
        save_project(project_name, source_filename, st.session_state.docs_edit, st.session_state.proc_edit, st.session_state.milestones_edit)
        st.success(f"Saved latest changes for project: {project_name}")

# =========================================================
# PROJECT DASHBOARD
# =========================================================
if page == "Project Dashboard":
    st.subheader(f"Project Dashboard — {project_name}")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Documents", len(docs_scored))
    c2.metric("Procurement", len(proc_scored))
    c3.metric("Overdue Docs", int((docs_scored["Document Status"] == "OVERDUE").sum()) if not docs_scored.empty else 0)
    c4.metric("Overdue Procurement", int((proc_scored["Overall Status"] == "OVERDUE").sum()) if not proc_scored.empty else 0)
    c5.metric("FAT Readiness", f"{readiness['Overall FAT Readiness %']}%")

    left, right = st.columns(2)
    with left:
        st.markdown("### Document status")
        if not docs_scored.empty:
            st.bar_chart(docs_scored["Document Status"].value_counts())
    with right:
        st.markdown("### Procurement status")
        if not proc_scored.empty:
            st.bar_chart(proc_scored["Overall Status"].value_counts())

# =========================================================
# DOCUMENTS
# =========================================================
elif page == "Documents Tracker":
    st.subheader(f"Documents Tracker — {project_name}")
    cols = ["Project", "Document No", "Document Title", "Expected Send to Client", "Current Forecast Send", "Sent to Client?", "Actual Sent Date", "Expected Return from Client", "Current Forecast Return", "Required for FAT?", "Delay Reason / Notes"]
    edited = st.data_editor(st.session_state.docs_edit[cols], use_container_width=True, hide_index=True, num_rows="dynamic") if not st.session_state.docs_edit.empty else pd.DataFrame(columns=cols)
    st.session_state.docs_edit = edited
    docs_scored = apply_document_status(st.session_state.docs_edit, assessment_date)
    st.markdown("### Highlighted status")
    st.dataframe(style_status(docs_scored, "Document Status"), use_container_width=True, hide_index=True)
    if st.button("Save document changes", type="primary"):
        save_current_project()

# =========================================================
# PROCUREMENT
# =========================================================
elif page == "Procurement Tracker":
    st.subheader(f"Procurement Tracker — {project_name}")
    search = st.text_input("Search item / document / dependency")
    cols = ["Project", "Item / Instrument", "Linked Document No", "Linked Document Title", "Expected RFQ Date", "RFQ Placed?", "Actual RFQ Date", "Expected PO/Order Date", "Ordered?", "Actual PO/Order Date", "Expected Delivery Date", "Received / Delivered?", "Baseline Build/Test", "Required for FAT?", "Dependency / Notes", "Delay Reason / Notes"]
    view = st.session_state.proc_edit[cols] if not st.session_state.proc_edit.empty else pd.DataFrame(columns=cols)
    if search and not view.empty:
        view = view[view.astype(str).apply(lambda row: row.str.contains(search, case=False, na=False).any(), axis=1)]
    edited = st.data_editor(view, use_container_width=True, hide_index=True, num_rows="dynamic")
    if not search:
        st.session_state.proc_edit = edited
    else:
        temp = st.session_state.proc_edit.copy()
        for idx in edited.index:
            temp.loc[idx, cols] = edited.loc[idx, cols]
        st.session_state.proc_edit = temp
    proc_scored = apply_procurement_status(st.session_state.proc_edit, assessment_date)
    st.markdown("### Highlighted status")
    st.dataframe(style_status(proc_scored, "Overall Status"), use_container_width=True, hide_index=True)
    if st.button("Save procurement changes", type="primary"):
        save_current_project()

# =========================================================
# FAT READINESS
# =========================================================
elif page == "FAT Readiness":
    st.subheader(f"FAT Readiness — {project_name}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Overall FAT Readiness", f"{readiness['Overall FAT Readiness %']}%")
    c2.metric("Document Readiness", f"{readiness['Document Readiness %']}%", readiness["FAT Documents Complete"])
    c3.metric("Procurement Readiness", f"{readiness['Procurement Readiness %']}%", readiness["FAT Procurement Complete"])
    c4.metric("Milestone Readiness", f"{readiness['FAT Milestone Readiness %']}%", readiness["FAT Milestones Complete"])

    st.dataframe(readiness_df, use_container_width=True, hide_index=True)

    st.markdown("### FAT-critical documents")
    fat_docs = docs_scored[docs_scored["Required for FAT?"] == True] if not docs_scored.empty else docs_scored
    st.dataframe(style_status(fat_docs, "Document Status"), use_container_width=True, hide_index=True)

    st.markdown("### FAT-critical procurement")
    fat_proc = proc_scored[proc_scored["Required for FAT?"] == True] if not proc_scored.empty else proc_scored
    st.dataframe(style_status(fat_proc, "Overall Status"), use_container_width=True, hide_index=True)

    st.markdown("### FAT / test milestone checklist")
    edited_ms = st.data_editor(st.session_state.milestones_edit, use_container_width=True, hide_index=True, num_rows="dynamic")
    st.session_state.milestones_edit = edited_ms
    if st.button("Save FAT readiness changes", type="primary"):
        save_current_project()

# =========================================================
# EMAIL REMINDERS
# =========================================================
elif page == "Email Reminders":
    st.subheader(f"Email Reminders — {project_name}")
    records = []
    if not docs_scored.empty:
        for _, row in docs_scored[docs_scored["Document Status"].isin(["OVERDUE", "DUE THIS WEEK"])].iterrows():
            subject = f"[{row['Document Status']}] Document reminder - {row['Document No']} {row['Document Title']}"
            body = f"""Hi {engineer_name},

Please review the following document submission:

Project: {project_name}
Document No: {row['Document No']}
Document Title: {row['Document Title']}
Expected Send to Client: {fmt_date(row.get('Expected Send to Client'))}
Current Forecast Send: {fmt_date(row.get('Current Forecast Send'))}
Status: {row['Document Status']}

Please update the tracker once actioned.

Regards,
E&I Tracker
"""
            records.append({"Type": "Document", "Status": row["Document Status"], "Item": f"{row['Document No']} - {row['Document Title']}", "Subject": subject, "Body": body, "Link": mailto_link(engineer_email, subject, body)})
    if not proc_scored.empty:
        for _, row in proc_scored[proc_scored["Overall Status"].isin(["OVERDUE", "DUE THIS WEEK"])].iterrows():
            subject = f"[{row['Overall Status']}] Procurement reminder - {row['Item / Instrument']}"
            body = f"""Hi {engineer_name},

Please review the following procurement item:

Project: {project_name}
Item / Instrument: {row['Item / Instrument']}
Expected RFQ Date: {fmt_date(row.get('Expected RFQ Date'))}
Expected PO / Order Date: {fmt_date(row.get('Expected PO/Order Date'))}
Expected Delivery Date: {fmt_date(row.get('Expected Delivery Date'))}
Status: {row['Overall Status']}
Dependency / Notes: {row.get('Dependency / Notes', '')}

Please update the tracker once actioned.

Regards,
E&I Tracker
"""
            records.append({"Type": "Procurement", "Status": row["Overall Status"], "Item": row["Item / Instrument"], "Subject": subject, "Body": body, "Link": mailto_link(engineer_email, subject, body)})

    if readiness["FAT Readiness Status"] != "READY":
        subject = f"[{readiness['FAT Readiness Status']}] FAT readiness review required - {project_name}"
        body = f"""Hi {engineer_name},

FAT readiness currently requires review for project {project_name}.

Overall FAT Readiness: {readiness['Overall FAT Readiness %']}%
Document Readiness: {readiness['Document Readiness %']}% ({readiness['FAT Documents Complete']})
Procurement Readiness: {readiness['Procurement Readiness %']}% ({readiness['FAT Procurement Complete']})
Milestone Readiness: {readiness['FAT Milestone Readiness %']}% ({readiness['FAT Milestones Complete']})

Please review outstanding FAT-critical documents, procurement items, and FAT/test milestones.

Regards,
E&I Tracker
"""
        records.append({"Type": "FAT Readiness", "Status": readiness["FAT Readiness Status"], "Item": "FAT readiness", "Subject": subject, "Body": body, "Link": mailto_link(engineer_email, subject, body)})

    reminders = pd.DataFrame(records)
    st.write(f"Reminders generated: **{len(reminders)}**")
    if reminders.empty:
        st.success("No reminders required.")
    else:
        for idx, row in reminders.iterrows():
            with st.expander(f"{row['Status']} - {row['Type']} - {row['Item']}"):
                st.text_area("Email body", row["Body"], height=220, key=f"email_{idx}")
                st.markdown(f"[Open Outlook draft]({row['Link']})")
        st.download_button("Download reminder email log", reminders.drop(columns=["Link"]).to_csv(index=False).encode("utf-8"), "SPOTT_Reminder_Email_Log.csv", "text/csv")

# =========================================================
# EXPORT / BACKUP
# =========================================================
elif page == "Export / Backup":
    st.subheader(f"Export / Backup — {project_name}")
    docs_out = apply_document_status(st.session_state.docs_edit, assessment_date)
    proc_out = apply_procurement_status(st.session_state.proc_edit, assessment_date)
    readiness_out = pd.DataFrame([calculate_fat_readiness(docs_out, proc_out, st.session_state.milestones_edit)])

    st.download_button(
        "Download selected project Excel",
        make_excel_download(docs_out, proc_out, st.session_state.milestones_edit, readiness_out),
        f"{project_name}_SPOTT_Tracker_Output.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button("Download complete database backup", export_database_json(), "spott_project_tracker_backup.json", "application/json")

    backup_upload = st.file_uploader("Restore database backup JSON", type=["json"])
    if st.button("Restore backup"):
        if not backup_upload:
            st.warning("Please upload a backup JSON file first.")
        else:
            import_database_json(backup_upload)
            st.success("Backup restored. Refresh or navigate to Project Library to view projects.")
